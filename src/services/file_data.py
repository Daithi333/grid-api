import io
import logging
import os
import tempfile
from datetime import datetime
from typing import List

from openpyxl import load_workbook, Workbook
from openpyxl.cell import ReadOnlyCell
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.worksheet import Worksheet

from constants import DATE_FORMAT, DATE_STYLE
from context import db_session
from database.models import File, Change
from decorators import enforce_permission
from services import file_cache, FileService
from enums import ChangeType
from util.subprocess import open_close_excel

logger = logging.getLogger(__name__)


class FileDataService:

    @classmethod
    @enforce_permission(file_id_key='id_', required_roles=['*'])
    def get_data(cls, id_: str) -> dict:
        file = FileService.get(id_=id_, internal=True)

        cells: List[List[ReadOnlyCell]] = file_cache.load_excel(file)
        data_types = file.data_types

        row_data = []
        for row_number, row_cells in enumerate(cells[1:]):
            row = {'_rowNumber': row_number + 1}
            for i, hc in enumerate(cells[0]):
                if data_types[hc.value] == 'd':
                    try:
                        cell_value = row_cells[i].value.strftime(DATE_FORMAT)
                    except AttributeError:
                        cell_value = row_cells[i].value
                else:
                    cell_value = row_cells[i].value

                row[hc.value] = cell_value if cell_value is not None else ''

            row_data.append(row)

        return {
            'columnDefs': cls._get_column_definitions(cells, file.data_types),
            'rowData': row_data
        }

    @classmethod
    def _get_column_definitions(cls, cells, data_types: dict) -> List[dict]:
        """Get ag-grid column properties based on openpyxl data_types"""

        data_type_props = {
            'n': {'filter': 'agNumberColumnFilter'},
            's': {'filter': 'agTextColumnFilter'},
            'd': {'filter': 'agDateColumnFilter'},
            'f': {'filter': 'agNumberColumnFilter', 'editable': False},
            'e': {'filter': 'agNumberColumnFilter', 'editable': False}
        }

        column_definitions = []
        for hc in cells[0]:
            data_type = data_types[hc.value]
            column_def = {'field': hc.value, 'colId': hc.column_letter, **data_type_props[data_type]}
            if data_type in ['e', 'f']:
                column_def['headerName'] = column_def['field'] + '*'
            column_definitions.append(column_def)

        return column_definitions

    @classmethod
    def get_data_types(cls, file_bytes: bytes) -> dict:
        """Return dictionary of column header to data type, based on the first populated row for each column."""
        wb = cls._load_workbook(file_bytes)
        ws = wb.active
        cells = list(ws.rows)[:]

        data_types = [cell.data_type if cell.value else None for cell in cells[1]]

        for col_index, data_type in enumerate(data_types):
            if data_type is not None:
                continue

            data_types[col_index] = 'n'  # default to 'n' as openpyxl does, in case entire column is empty
            for row_cells in cells[2:]:
                if row_cells[col_index].value is not None:
                    data_types[col_index] = row_cells[col_index].data_type
                    break

        return {hc.value: dt for hc, dt in zip(cells[0], data_types)}

    @classmethod
    def apply_changes(cls, file: File, changes: List[Change]) -> bool:
        """
        Apply transaction changes to the actual saved excel file.
        Changes are applied in reverse row number order, so creates and deletes don't affect subsequent changes.
        TODO - Other pending transactions will still be affected and need a solution.
        """
        logger.info('Apply changes to workbook - begin')
        wb = cls._load_workbook(file.blob)
        ws = wb.active  # only get single (first) worksheet for now

        # take a snapshot of row2 values for use in formula translation
        value_lookup = {cell.coordinate: cell.value for cell in list(ws[2])}

        changes.sort(key=lambda x: x.row_number, reverse=True)
        for change in changes:
            if change.change_type == ChangeType.CREATE:
                cls._handle_create(ws, change, file.data_types)
            elif change.change_type == ChangeType.UPDATE:
                cls._handle_update(ws, change)
            else:
                cls._handle_delete(ws, change)

        cls._regenerate_formulas(ws, file.data_types, value_lookup)

        session = db_session.get()
        file_bytes = cls._convert_to_bytes(wb, file.name)
        logger.info('Converted workbook to bytes')
        file.blob = file_bytes
        session.commit()
        file_cache.remove(file.id)  # remove old version of file data from cache
        logger.info('Apply changes to workbook - complete')
        return True

    @classmethod
    def _load_workbook(cls, file_bytes: bytes) -> Workbook:
        """Load workbook in read/write mode for updating contents or getting data types"""
        virtual_file = io.BytesIO(file_bytes)
        wb = load_workbook(virtual_file, read_only=False, data_only=False)
        return wb

    @classmethod
    def _handle_create(cls, ws: Worksheet, change: Change, data_types: dict):
        """Add a new row to the end of a spreadsheet"""
        logger.info('Create new row - begin')
        new_row_number = ws.max_row + 1
        new_row_data = change.after
        header_cells = list(ws[1])

        for hc in header_cells:
            if hc.value not in new_row_data:
                raise Exception(f'Column {hc.value!r} not found in new row data for Change {change.id!r}')

            data_type = data_types[hc.value]
            new_cell = ws[f'{hc.column_letter}{new_row_number}']
            new_value_str = new_row_data[hc.value]

            if new_value_str is None:
                continue

            if data_type in ['e', 'f']:
                pass  # formulas regenerated at end of the transaction
            elif data_type == 'd':
                new_cell.value = datetime.strptime(new_value_str, DATE_FORMAT)
                new_cell.number_format = DATE_STYLE
            elif data_type == 'n':
                if new_value_str.isdigit():
                    new_value = int(new_value_str)
                else:
                    try:
                        new_value = float(new_value_str)
                    except ValueError:
                        new_value = new_value_str
                new_cell.value = new_value
            else:
                new_cell.value = new_value_str
        logger.info('Create new row - complete')

    @classmethod
    def _handle_delete(cls, ws: Worksheet, change: Change):
        """
        Delete row specified by change row_number, if data matches change before value
        If data does not match, iterate backwards to find a matching row to delete.
        TODO - Row insertions not supported yet, but logic would need to change to support those.
        """
        logger.info('Delete row - begin')
        deletion_row_number = change.row_number + 1
        header_cells = list(ws[1])
        row_cells = list(ws[deletion_row_number])

        if cls._is_match(header_cells, row_cells, change.before):
            ws.delete_rows(deletion_row_number)
        else:
            logger.warning('Unable to delete row as it is not as expected')
            for alt_deletion_row_number in range(deletion_row_number-1, 2, -1):
                row_cells = list(ws[alt_deletion_row_number])
                if cls._is_match(header_cells, row_cells, change.before):
                    ws.delete_rows(alt_deletion_row_number)
                    logger.warning(f'Deleted row {alt_deletion_row_number} instead of {deletion_row_number}')
                    break
        logger.info('Delete row - complete')

    @classmethod
    def _is_match(cls, header_cells, row_cells, change_before) -> bool:
        for hc, rc in zip(header_cells, row_cells):
            excel_value = rc.value
            if excel_value is None and change_before[hc.value] in ['', None]:
                continue
            if rc.data_type in ['e', 'f']:
                continue  # ignore value check on formula cells
            if rc.data_type == 'n':
                excel_value = str(excel_value)
            if rc.data_type == 'd':
                excel_value = excel_value.strftime(DATE_FORMAT)
            if excel_value != change_before[hc.value]:
                return False
        return True

    @staticmethod
    def _handle_update(ws: Worksheet, change: Change):
        """Update an existing row in the spreadsheet"""
        logger.info('Update new row - begin')
        update_row_number = change.row_number + 1
        header_cells = list(ws[1])
        row_cells = list(ws[update_row_number])

        for hc, rc in zip(header_cells, row_cells):
            if hc.value not in change.before:
                raise Exception(f'Column {hc.value!r} not found in update row data for Change {change.id!r}')

            updated_value_str = str(change.after[hc.value])
            if rc.data_type in ['e', 'f']:
                continue
            elif rc.data_type == 'n':
                if updated_value_str.isdigit():
                    updated_value = int(updated_value_str)
                else:
                    try:
                        updated_value = float(updated_value_str)
                    except ValueError:
                        updated_value = updated_value_str
                rc.value = updated_value
            elif rc.data_type == 'd':
                rc.value = datetime.strptime(updated_value_str, DATE_FORMAT)
                rc.number_format = DATE_STYLE
            else:
                rc.value = updated_value_str
        logger.info('Update new row - complete')

    @classmethod
    def _regenerate_formulas(cls, ws: Worksheet, data_types: dict, value_lookup: dict):
        logger.info('Regenerate cell formulas - begin')
        for hc in list(ws[1]):
            if data_types[hc.value] in ['e', 'f']:
                column_letter = hc.column_letter
                origin_coord = f'{column_letter}2'
                origin_formula = value_lookup[origin_coord]
                column_cells = list(ws[column_letter])
                for i, cell in enumerate(column_cells[1:]):
                    destination_coord = f'{column_letter}{i+2}'
                    cell.value = Translator(origin_formula, origin=origin_coord).translate_formula(destination_coord)
        logger.info('Regenerate cell formulas - complete')

    @classmethod
    def _convert_to_bytes(cls, wb: Workbook, filename: str):
        """Convert workbook to bytes for saving into database. Will attempt to open-close the excel to evaluate
        the formulas and cache the results, which differs per OS and requires MS Excel or LibreOffice installation.
        Future alternative could be to use a python library to evaluate all the formulas when data is being returned.
        """
        logger.info('Converting workbook to bytes')
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = os.path.join(temp_dir, f'temp-{filename}')
            wb.save(temp_path)
            open_close_excel(temp_path)
            with open(temp_path, 'rb') as temp_file:
                return temp_file.read()
