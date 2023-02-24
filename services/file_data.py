import io
from typing import List

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from database import db
from database.models import File, Change
from services import file_cache
from services.file_cache import load_excel
from enums import ChangeType


class FileData:

    @classmethod
    def get_data(cls, file: File):
        cells = load_excel(file)

        row_data = []
        for row_number, row_cells in enumerate(cells[1:]):
            row = {'_rowNumber': row_number + 1}
            for i, hc in enumerate(cells[0]):
                row[hc.value] = row_cells[i].value
            row_data.append(row)

        return {
            'columnDefs': cls._get_column_definitions(cells, file.data_types),
            'rowData': row_data
        }

    @classmethod
    def _get_column_definitions(cls, cells, data_types: dict) -> List[dict]:
        """Get ag-grid column properties based on openpyxl data_types"""

        data_type_props = {
            'n': {'filter': 'agNumberColumnFilter'},
            's': {'filter': 'agTextColumnFilter'},
            'd': {'filter': 'agDateColumnFilter'},
            'f': {'filter': 'agNumberColumnFilter', 'editable': False},
            'e': {'filter': 'agNumberColumnFilter', 'editable': False}
        }

        column_definitions = []
        for hc in cells[0]:
            data_type = data_types[hc.value]
            column_def = {'field': hc.value, 'colId': hc.column_letter, **data_type_props[data_type]}
            if data_type in ['e', 'f']:
                column_def['headerName'] = column_def['field'] + '*'
            column_definitions.append(column_def)

        return column_definitions

    @classmethod
    def get_data_types(cls, file_bytes: bytes) -> dict:
        """Return dictionary of column header to data type, based on the first populated row for each column."""
        wb = cls._load_workbook(file_bytes)
        ws = wb.active
        cells = list(ws.rows)[:]

        data_types = [cell.data_type if cell.value else None for cell in cells[1]]

        for col_index, data_type in enumerate(data_types):
            if data_type is not None:
                continue

            data_types[col_index] = 'n'  # default to 'n' as openpyxl does, in case entire column is empty
            for row_cells in cells[2:]:
                if row_cells[col_index].value is not None:
                    data_types[col_index] = row_cells[col_index].data_type
                    break

        return {hc.value: dt for hc, dt in zip(cells[0], data_types)}

    @classmethod
    def apply_changes(cls, file: File, changes: List[Change]) -> bool:
        """Apply transaction changes to the actual saved excel file.
           Changes are applied in reverse row number order, so creates and deletes don't affect subsequent changes.
           TODO - Other pending transactions will still be affected and need a solution.
        """
        wb = cls._load_workbook(file.blob)
        ws = wb.active  # only get single (first) worksheet for now

        changes.sort(key=lambda x: x.row_number, reverse=True)
        for change in changes:
            if change.change_type == ChangeType.CREATE:
                cls._handle_create(ws, change)
            elif change.change_type == ChangeType.DELETE:
                cls._handle_delete(ws, change)
            else:  # update
                cls._handle_update(ws, change)

        session = next(db.get_session())
        wb_bytes = io.BytesIO()
        wb.save(wb_bytes)
        file.blob = wb_bytes.getvalue()
        session.commit()
        file_cache.remove(file.id)  # remove old version of file data from cache
        return True

    @classmethod
    def _load_workbook(cls, file_bytes: bytes) -> Workbook:
        """Load workbook in read/write mode for updating contents or getting data types"""
        virtual_file = io.BytesIO(file_bytes)
        wb = load_workbook(virtual_file, read_only=False, data_only=False)
        return wb

    @staticmethod
    def _handle_create(ws: Worksheet, change: Change):
        """Add a new row to the end of a spreadsheet"""
        new_row_number = ws.max_row + 1
        new_row_data = change.after
        header_cells = list(ws[1])

        for hc in header_cells:
            if hc.value not in new_row_data:
                raise Exception(f'Column {hc.value!r} not found in new row data for Change {change.id!r}')
            new_value = new_row_data[hc.value]
            ws[f'{hc.column_letter}{new_row_number}'] = new_value
            # TODO - handle formula columns

    @staticmethod
    def _handle_delete(ws: Worksheet, change: Change):
        """Delete row specified by change row_number, if data matches change before value"""
        deletion_row_number = change.row_number + 1
        header_cells = list(ws[1])
        row_cells = list(ws[deletion_row_number])

        mismatches = []
        for hc, rc in zip(header_cells, row_cells):
            if rc.value != change.before[hc.value]:
                mismatches.append({'header': hc.value, 'row value': rc.value, 'change': change.before[hc.value]})

        if mismatches:
            print(mismatches)
            raise Exception(f'Worksheet row to be deleted does not match change {change.id!r} deletion data')

        ws.delete_rows(deletion_row_number)

    @staticmethod
    def _handle_update(ws: Worksheet, change: Change):
        """Update an existing row in the spreadsheet"""
        update_row_number = change.row_number + 1
        header_cells = list(ws[1])

        for hc in header_cells:
            if hc.value not in change.before:
                raise Exception(f'Column {hc.value!r} not found in update row data for Change {change.id!r}')
            new_value = change.after[hc.value]
            ws[f'{hc.column_letter}{update_row_number}'] = new_value
            # TODO - handle formula columns
